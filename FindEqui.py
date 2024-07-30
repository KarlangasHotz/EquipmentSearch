import os
from openpyxl import load_workbook
import pandas as pd
from tkinter import Tk, Label, Entry, Button, Text, Scrollbar, messagebox, INSERT, END
import threading
from tkinter import ttk
from concurrent.futures import ThreadPoolExecutor, as_completed

# Flag to monitor the status of the search thread
search_thread_running = False

def perform_search():
    global search_thread_running
    equipment_numbers = entry_equipment_numbers.get().split(',')
    data = []
    mainPath = "C:/"

    def process_file(file_path):
        try:
            wb = load_workbook(file_path, read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
                    if row[0] in equipment_numbers:
                        equipment_number = row[0]
                        relative_path = os.path.relpath(os.path.dirname(file_path), start=mainPath)
                        result_text = f"'{equipment_number}' was found in '{relative_path}'"
                        text_output.insert(INSERT, result_text + '\n')
                        data.append((equipment_number, os.path.basename(file_path), relative_path))
        except Exception as e:
            error_text = f"Error processing file '{os.path.basename(file_path)}': {str(e)}"
            text_output.insert(INSERT, error_text + '\n')

    # Using ThreadPoolExecutor to parallelize file processing
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = []
        for root, _, files in os.walk(mainPath):
            for filename in files:
                if filename.endswith(".xlsx"):
                    file_path = os.path.join(root, filename)
                    futures.append(executor.submit(process_file, file_path))
        for future in as_completed(futures):
            future.result()  # Ensures exceptions are raised in the main thread

    df = pd.DataFrame(data, columns=["Equipment Number", "File Name", "Relative Path"])
    results_file_path = os.path.join(os.path.dirname(__file__), "Results.xlsx")
    df.to_excel(results_file_path, index=False)
    messagebox.showinfo("Results", f"Results saved to '{results_file_path}'")

    # Reset the flag when the search is completed
    search_thread_running = False

def search_equipment():
    global search_thread_running
    if search_thread_running:
        messagebox.showinfo("Info", "Search is already running.")
        return

    label_result.config(text="")
    text_output.delete(1.0, END)
    search_thread_running = True
    search_thread = threading.Thread(target=perform_search)
    search_thread.start()
    progress_bar.start()

    def monitor_thread():
        global search_thread_running
        while search_thread_running:
            if not search_thread.is_alive():
                progress_bar.stop()
                messagebox.showinfo("Info", "Search stopped.")
                break
    threading.Thread(target=monitor_thread).start()

root = Tk()
root.title("Equipment Search")
label_instruction = Label(root, text="Enter Equipment Numbers (comma-separated):")
label_instruction.pack()
entry_equipment_numbers = Entry(root)
entry_equipment_numbers.pack()
button_search = Button(root, text="Search Equipment", command=search_equipment)
button_search.pack()
label_result = Label(root, text="Search Results:")
label_result.pack()
text_output = Text(root, height=10, width=80, wrap="none")
text_output.pack(side="left", fill="both", expand=True)
scrollbar = Scrollbar(root, command=text_output.yview)
scrollbar.pack(side="right", fill="y")
text_output.configure(yscrollcommand=scrollbar.set)
progress_bar = ttk.Progressbar(root, mode="indeterminate")
progress_bar.pack(fill="x")
root.mainloop()
